#!/usr/bin/env python3
"""
Calculator Generator Agent.

Generates client-facing Excel calculator spreadsheets and uploads
to the "Cost Calculators" Team Folder in Zoho WorkDrive.
"""

import json
import os
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

from zoho_config import ZOHO_QUOTE_DEFAULTS

# Import JobStateManager for state updates (optional dependency)
try:
    from job_state import JobStateManager, WorkflowStatus
except ImportError:
    JobStateManager = None
    WorkflowStatus = None


# =============================================================================
# Result Types
# =============================================================================

@dataclass
class CalculatorResult:
    """Result from calculator generation."""
    success: bool
    file_path: Optional[str] = None
    file_name: Optional[str] = None
    drive_file_id: Optional[str] = None
    drive_permalink: Optional[str] = None
    products_count: int = 0
    error: Optional[str] = None
    duration_seconds: float = 0


# =============================================================================
# Tool Definitions (simplified - 3 tools)
# =============================================================================

CALCULATOR_AGENT_TOOLS = [
    {
        "name": "generate_calculator_xlsx",
        "description": "Generate Excel calculator spreadsheet. Returns local file path.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_name": {
                    "type": "string",
                    "description": "File name without extension (e.g., 'Otava Promo Calc 2026')"
                }
            },
            "required": ["file_name"]
        }
    },
    {
        "name": "upload_to_cost_calculators",
        "description": "Upload the Excel file to the 'Cost Calculators' Team Folder in Zoho WorkDrive.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Local path to the Excel file"
                }
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "report_completion",
        "description": "Report that calculator generation and upload is complete.",
        "input_schema": {
            "type": "object",
            "properties": {
                "summary": {"type": "string"},
                "file_name": {"type": "string"},
                "permalink": {"type": "string"}
            },
            "required": ["summary"]
        }
    }
]


# =============================================================================
# System Prompt
# =============================================================================

CALCULATOR_AGENT_SYSTEM_PROMPT = """You are a Calculator Generator agent.

## Your Task
1. Generate an Excel calculator spreadsheet for the client
2. Upload it to the "Cost Calculators" folder in Zoho WorkDrive
3. Report completion with the permalink

## Workflow

STEP 1: Call generate_calculator_xlsx
- File name format: "{ClientName} Promo Calc {YYYYMMDD_HHMMSS}"
- Example: "Otava Promo Calc 20251213_143022"
- This ensures each calculator is saved as a new file, not overwriting existing ones
- Clean the client name (remove special characters)

STEP 2: Call upload_to_cost_calculators
- Pass the file_path returned from step 1

STEP 3: Call report_completion
- Include the permalink so user can access the file

## Notes
- The spreadsheet has editable quantity columns (yellow) for client to adjust
- Formulas auto-calculate totals and freight estimates
- Keep file names descriptive but concise
"""


# =============================================================================
# Calculator Generator Agent Class
# =============================================================================

class CalculatorGeneratorAgent:
    """Agent that generates Excel calculators and uploads to Zoho WorkDrive."""

    def __init__(
        self,
        zoho_client=None,
        anthropic_client=None,
        model: str = "claude-sonnet-4-20250514",
        max_tokens: int = 4096,
        max_iterations: int = 10,
        state_manager: Optional["JobStateManager"] = None
    ):
        self.zoho_client = zoho_client
        self.anthropic = anthropic_client or anthropic.Anthropic()
        self.model = model
        self.max_tokens = max_tokens
        self.max_iterations = max_iterations
        self.state_manager = state_manager

        # State
        self._unified_output: Dict[str, Any] = {}
        self._result: Optional[CalculatorResult] = None
        self._agent_complete: bool = False
        self._generated_file_path: Optional[str] = None
        self._output_dir: str = "./output"

    def _update_state(self, status: str, **kwargs) -> None:
        """Update job state if state manager is available."""
        if self.state_manager and WorkflowStatus:
            self.state_manager.update(status, **kwargs)

    def generate_calculator(
        self,
        unified_output: Dict[str, Any],
        output_dir: str = "./output",
        dry_run: bool = False
    ) -> CalculatorResult:
        """
        Main entry point - generate calculator and optionally upload to Drive.

        Args:
            unified_output: The unified output from ESP/SAGE pipeline
            output_dir: Directory to save the generated Excel file
            dry_run: If True, only generate locally without uploading

        Returns:
            CalculatorResult with success status and file details
        """
        start_time = time.time()

        self._unified_output = unified_output
        self._output_dir = output_dir
        self._agent_complete = False
        self._result = None
        self._generated_file_path = None

        # Extract client info
        client_name = unified_output.get("client", {}).get("company") or \
                      unified_output.get("client", {}).get("name") or "Client"
        products = unified_output.get("products", [])

        # Build initial message with context
        initial_message = f"""Generate a calculator spreadsheet for:

Client: {client_name}
Products: {len(products)} items
Year: {datetime.now().year}

Product summary:
"""
        for i, p in enumerate(products[:5], 1):  # Show first 5
            name = p.get("item", {}).get("name", "Unknown")
            price_breaks = p.get("pricing", {}).get("breaks", [])
            price = price_breaks[0].get("sell_price", 0) if price_breaks else 0
            initial_message += f"  {i}. {name} (${price:.2f})\n"
        if len(products) > 5:
            initial_message += f"  ... and {len(products) - 5} more\n"

        initial_message += "\nPlease generate the calculator and upload to Cost Calculators folder."

        # Dry run mode - just generate locally
        if dry_run:
            # Clean client name for filename
            clean_name = "".join(c for c in client_name if c.isalnum() or c in " -_").strip()
            file_name = f"{clean_name} Promo Calc {datetime.now().strftime('%Y%m%d_%H%M%S')}"

            try:
                file_path = self._generate_xlsx(file_name)
                return CalculatorResult(
                    success=True,
                    file_path=file_path,
                    file_name=f"{file_name}.xlsx",
                    products_count=len(products),
                    duration_seconds=time.time() - start_time
                )
            except Exception as e:
                return CalculatorResult(
                    success=False,
                    error=str(e),
                    duration_seconds=time.time() - start_time
                )

        # Run agent loop for full operation
        messages = [{"role": "user", "content": initial_message}]

        for iteration in range(self.max_iterations):
            if self._agent_complete:
                break

            try:
                response = self.anthropic.messages.create(
                    model=self.model,
                    max_tokens=self.max_tokens,
                    system=CALCULATOR_AGENT_SYSTEM_PROMPT,
                    tools=CALCULATOR_AGENT_TOOLS,
                    messages=messages
                )
            except Exception as e:
                return CalculatorResult(
                    success=False,
                    error=f"Anthropic API error: {e}",
                    duration_seconds=time.time() - start_time
                )

            # Process response
            assistant_content = []
            tool_results = []

            for block in response.content:
                if block.type == "text":
                    assistant_content.append({"type": "text", "text": block.text})
                elif block.type == "tool_use":
                    assistant_content.append({
                        "type": "tool_use",
                        "id": block.id,
                        "name": block.name,
                        "input": block.input
                    })
                    # Execute tool
                    result = self._handle_tool_call(block.name, block.input)
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result
                    })

            messages.append({"role": "assistant", "content": assistant_content})
            if tool_results:
                messages.append({"role": "user", "content": tool_results})

            if response.stop_reason == "end_turn" and not tool_results:
                break

        if self._result:
            self._result.duration_seconds = time.time() - start_time
            return self._result

        return CalculatorResult(
            success=False,
            error="Agent did not complete",
            duration_seconds=time.time() - start_time
        )

    def _handle_tool_call(self, tool_name: str, tool_input: Dict) -> str:
        """Dispatch tool calls to appropriate handlers."""
        if tool_name == "generate_calculator_xlsx":
            return self._tool_generate_xlsx(tool_input)
        elif tool_name == "upload_to_cost_calculators":
            return self._tool_upload(tool_input)
        elif tool_name == "report_completion":
            return self._tool_report_completion(tool_input)
        return json.dumps({"error": f"Unknown tool: {tool_name}"})

    def _tool_generate_xlsx(self, tool_input: Dict) -> str:
        """Tool handler: Generate the Excel spreadsheet."""
        # Emit state: generating calculator
        self._update_state(
            WorkflowStatus.CALC_GENERATING.value if WorkflowStatus else "calc_generating"
        )

        # Emit thought for generating calculator
        if self.state_manager:
            self.state_manager.emit_thought(
                agent="calculator_agent",
                event_type="action",
                content="Generating cost calculator spreadsheet"
            )

        file_name = tool_input.get("file_name", "Calculator")
        # Clean filename
        file_name = "".join(c for c in file_name if c.isalnum() or c in " -_").strip()

        try:
            file_path = self._generate_xlsx(file_name)
            self._generated_file_path = file_path
            return json.dumps({
                "success": True,
                "file_path": file_path,
                "file_name": f"{file_name}.xlsx",
                "products_count": len(self._unified_output.get("products", []))
            })
        except Exception as e:
            return json.dumps({"success": False, "error": str(e)})

    def _tool_upload(self, tool_input: Dict) -> str:
        """Tool handler: Upload to Cost Calculators folder."""
        file_path = tool_input.get("file_path")
        if not file_path or not os.path.exists(file_path):
            return json.dumps({"success": False, "error": "File not found"})

        if not self.zoho_client:
            return json.dumps({"success": False, "error": "Zoho client not configured"})

        # Emit state: uploading calculator
        self._update_state(
            WorkflowStatus.CALC_UPLOADING.value if WorkflowStatus else "calc_uploading"
        )

        # Emit thought for uploading calculator
        if self.state_manager:
            self.state_manager.emit_thought(
                agent="calculator_agent",
                event_type="action",
                content="Uploading calculator to Zoho WorkDrive"
            )

        try:
            result = self.zoho_client.upload_to_cost_calculators(file_path)
            file_id = result.get("id")
            permalink = result.get("attributes", {}).get("permalink")

            return json.dumps({
                "success": True,
                "file_id": file_id,
                "permalink": permalink
            })
        except Exception as e:
            return json.dumps({"success": False, "error": str(e)})

    def _tool_report_completion(self, tool_input: Dict) -> str:
        """Tool handler: Mark agent as complete and record result."""
        self._agent_complete = True
        self._result = CalculatorResult(
            success=True,
            file_path=self._generated_file_path,
            file_name=tool_input.get("file_name"),
            drive_permalink=tool_input.get("permalink"),
            products_count=len(self._unified_output.get("products", []))
        )

        # Emit completion thought
        if self.state_manager:
            self.state_manager.emit_thought(
                agent="calculator_agent",
                event_type="checkpoint",
                content=f"Calculator complete: {tool_input.get('file_name')}",
                metadata={"permalink": tool_input.get("permalink")}
            )

        return json.dumps({"success": True, "message": "Completion recorded"})

    def _build_price_formula(self, breaks: List[Dict], qty_cell_ref: str) -> str:
        """
        Build a nested IF formula for quantity-based pricing.

        The formula checks quantities from highest to lowest tier,
        returning the appropriate price per unit.

        Example output: =IF(C3>=300,58.8,IF(C3>=204,61.12,IF(C3>=108,64.42,IF(C3>=48,69,75.16))))

        Args:
            breaks: List of price breaks with 'quantity' and 'sell_price'
            qty_cell_ref: Cell reference for quantity (e.g., "C3")

        Returns:
            Excel formula string
        """
        if not breaks:
            return "0"

        # Sort breaks by quantity descending (highest first for IF nesting)
        sorted_breaks = sorted(breaks, key=lambda b: b.get("quantity", 0) or 0, reverse=True)

        # Filter out breaks with no price
        valid_breaks = [b for b in sorted_breaks if b.get("sell_price") is not None]

        if not valid_breaks:
            return "0"

        # If only one price break, return the price directly
        if len(valid_breaks) == 1:
            return str(valid_breaks[0].get("sell_price", 0) or 0)

        # Build nested IF formula from highest quantity to lowest
        # Start with the lowest tier price as the default (innermost)
        formula = str(valid_breaks[-1].get("sell_price", 0) or 0)

        # Build nested IFs from second-lowest to highest
        for brk in reversed(valid_breaks[:-1]):
            qty = brk.get("quantity", 0) or 0
            price = brk.get("sell_price", 0) or 0
            formula = f"IF({qty_cell_ref}>={qty},{price},{formula})"

        return f"={formula}"

    def _generate_xlsx(self, file_name: str) -> str:
        """
        Generate Excel calculator spreadsheet using openpyxl.

        Features:
        - Min QTY column showing minimum order quantity
        - Your QTY column (editable) for client to input desired quantity
        - Price/Unit column with dynamic formula based on quantity tiers
        - Setup fee column
        - Total column (Price * Qty + Setup)
        - Freight estimate column

        Args:
            file_name: Name for the file (without extension)

        Returns:
            Path to the generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Calculator"

        # Styles
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Row 1: Instructions
        ws.merge_cells('A1:G1')
        ws['A1'] = "Enter your desired quantity in 'Your QTY' column - Price adjusts automatically based on quantity tiers"
        ws['A1'].fill = yellow_fill
        ws['A1'].font = bold_font

        # Row 2: Headers (7 columns now)
        headers = ["Promo Product", "Min QTY", "Your QTY", "Price/Unit", "Setup", "Total", "Freight Est."]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.border = thin_border

        # Get shipping percent from config
        shipping_percent = ZOHO_QUOTE_DEFAULTS.get("default_shipping_percent", 0.15)

        # Data rows
        products = self._unified_output.get("products", [])

        for row_idx, product in enumerate(products, 3):
            item = product.get("item", {})
            pricing = product.get("pricing", {})
            breaks = pricing.get("breaks", [])
            fees = product.get("fees", [])

            # Get first price break for min qty
            first_break = breaks[0] if breaks else {}
            min_qty = first_break.get("quantity", 0) or 0

            # Get setup fee
            setup_fee = 0
            for fee in fees:
                if fee.get("fee_type") == "setup":
                    setup_fee = fee.get("list_price", 0) or 0
                    break

            # Column A: Product Name
            name_cell = ws.cell(row=row_idx, column=1, value=item.get("name", ""))
            name_cell.border = thin_border

            # Column B: Min QTY (reference, not editable)
            min_qty_cell = ws.cell(row=row_idx, column=2, value=min_qty)
            min_qty_cell.border = thin_border
            min_qty_cell.fill = light_yellow_fill  # Light yellow to show it's reference

            # Column C: Your QTY (editable by client)
            your_qty_cell = ws.cell(row=row_idx, column=3, value=min_qty)  # Default to min
            your_qty_cell.fill = yellow_fill  # Bright yellow = editable
            your_qty_cell.border = thin_border
            your_qty_cell.font = bold_font

            # Column D: Price/Unit (dynamic formula based on quantity)
            qty_cell_ref = f"C{row_idx}"
            price_formula = self._build_price_formula(breaks, qty_cell_ref)
            price_cell = ws.cell(row=row_idx, column=4, value=price_formula)
            price_cell.border = thin_border
            price_cell.number_format = '$#,##0.00'

            # Column E: Setup
            setup_cell = ws.cell(row=row_idx, column=5, value=setup_fee)
            setup_cell.border = thin_border
            setup_cell.number_format = '$#,##0.00'

            # Column F: Total formula (Price * Qty + Setup)
            total_cell = ws.cell(row=row_idx, column=6, value=f"=D{row_idx}*C{row_idx}+E{row_idx}")
            total_cell.border = thin_border
            total_cell.number_format = '$#,##0.00'

            # Column G: Freight formula
            if shipping_percent > 0:
                freight_cell = ws.cell(row=row_idx, column=7, value=f"=F{row_idx}*{shipping_percent}")
                freight_cell.border = thin_border
                freight_cell.number_format = '$#,##0.00'
            else:
                freight_cell = ws.cell(row=row_idx, column=7, value="Included")
                freight_cell.border = thin_border

        # Totals row
        total_row = len(products) + 3

        # Label
        ws.cell(row=total_row, column=5, value="TOTALS:").font = bold_font

        # Grand total
        grand_total_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F3:F{total_row-1})")
        grand_total_cell.fill = green_fill
        grand_total_cell.font = bold_font
        grand_total_cell.number_format = '$#,##0.00'
        grand_total_cell.border = thin_border

        # Freight total
        freight_total_cell = ws.cell(row=total_row, column=7, value=f"=SUM(G3:G{total_row-1})")
        freight_total_cell.fill = green_fill
        freight_total_cell.font = bold_font
        freight_total_cell.number_format = '$#,##0.00'
        freight_total_cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

        # Add Price Breaks reference sheet
        self._add_price_breaks_sheet(wb, products)

        # Ensure output directory exists
        os.makedirs(self._output_dir, exist_ok=True)

        # Save
        file_path = os.path.join(self._output_dir, f"{file_name}.xlsx")
        wb.save(file_path)
        return file_path

    def _add_price_breaks_sheet(self, wb: Workbook, products: List[Dict]) -> None:
        """
        Add a reference sheet showing all price break tiers for each product.

        This helps the client understand the quantity thresholds and pricing.

        Args:
            wb: The workbook to add the sheet to
            products: List of products from unified output
        """
        ws = wb.create_sheet("Price Breaks Reference")

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        headers = ["Product", "Quantity", "Price/Unit", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        row_idx = 2
        for product in products:
            item = product.get("item", {})
            product_name = item.get("name", "Unknown")
            breaks = product.get("pricing", {}).get("breaks", [])
            price_includes = product.get("pricing", {}).get("price_includes", "")

            if not breaks:
                continue

            # Sort by quantity ascending for display
            sorted_breaks = sorted(breaks, key=lambda b: b.get("quantity", 0) or 0)

            for i, brk in enumerate(sorted_breaks):
                qty = brk.get("quantity", 0) or 0
                price = brk.get("sell_price", 0) or 0

                # Product name (only on first row for this product)
                name_cell = ws.cell(row=row_idx, column=1, value=product_name if i == 0 else "")
                name_cell.border = thin_border

                # Quantity
                qty_cell = ws.cell(row=row_idx, column=2, value=f"{qty}+")
                qty_cell.border = thin_border

                # Price
                price_cell = ws.cell(row=row_idx, column=3, value=price)
                price_cell.border = thin_border
                price_cell.number_format = '$#,##0.00'

                # Notes (only on first row)
                notes_cell = ws.cell(row=row_idx, column=4, value=price_includes if i == 0 else "")
                notes_cell.border = thin_border

                # Alternate row coloring
                if row_idx % 2 == 0:
                    for col in range(1, 5):
                        ws.cell(row=row_idx, column=col).fill = alt_fill

                row_idx += 1

            # Add blank row between products
            row_idx += 1

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30


# =============================================================================
# CLI Entry Point
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate calculator spreadsheet from unified output")
    parser.add_argument("input_file", help="Path to unified output JSON file")
    parser.add_argument("--output-dir", "-o", default="./output", help="Output directory")
    parser.add_argument("--dry-run", action="store_true", help="Generate locally without uploading")
    args = parser.parse_args()

    # Load unified output
    with open(args.input_file, "r") as f:
        unified_output = json.load(f)

    # Create agent and generate
    agent = CalculatorGeneratorAgent()
    result = agent.generate_calculator(
        unified_output=unified_output,
        output_dir=args.output_dir,
        dry_run=args.dry_run
    )

    if result.success:
        print(f"Calculator generated successfully!")
        print(f"  File: {result.file_path}")
        print(f"  Products: {result.products_count}")
        if result.drive_permalink:
            print(f"  Drive link: {result.drive_permalink}")
    else:
        print(f"Calculator generation failed: {result.error}")
